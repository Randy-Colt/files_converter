import re

from django.db import transaction
from django.db.models import Model
from openpyxl import load_workbook

from .exceptions import ExelParsingError
from .models import Album, Customer, Object


class AlbumExelParser:

    def __init__(self, file):
        self.file = file
        self.errors = []

    def validate_and_transform_volume(
        self,
        string_value: str | int | float | None,
        column: str,
        row_index: int
    ) -> float | None:
        try:
            if string_value is None or string_value == '':
                return None
            if isinstance(string_value, (float, int)):
                return float(string_value)
            return float(string_value.replace(',', '.'))
        except (ValueError, TypeError):
            self.errors.append(
                {f'{column}{row_index}': 'В ячейке не цифровой символ'}
            )

    def get_validate_inventory_num(
        self,
        filename: str,
        column: str,
        row_index: int
    ) -> str | None:
        if filename is None:
            return None
        match = re.search(r'(М-\d{6})', filename)
        if match:
            return match[0]
        self.errors.append(
            {
                f'{column}{row_index}':
                'Инвентарный номер в названии файла некорректен'
            }
        )

    def validate_name(
        self,
        value: str,
        column: str,
        row_index: int
    ) -> str | None:
        if value:
            return value
        self.errors.append(
            {f'{column}{row_index}': 'Поле не должно быть пустым'}
        )

    def validate_doc_type(
        self,
        value: str,
        column: str,
        row_index: int
    ) -> str | None:
        doc_type = Album.doc_mapping.get(value)
        if doc_type is not None:
            return doc_type
        self.errors.append(
            {f'{column}{row_index}': 'Неизвестный вид документации'}
        )

    def get_data_from_file(
        self
    ) -> tuple[list[dict[str, str | float], set[str], set[str]]]:
        worksheet = load_workbook(self.file, read_only=True).active
        albums = []
        customers_names = set()
        objects_names = set()
        iter_rows = enumerate(
            worksheet.iter_rows(2, values_only=True), start=2
        )

        for row_index, row in iter_rows:
            customer_name, obj_name, doc_type, volume, name, filename = row[:6]
            albums.append(
                {
                    'customer': self.validate_name(
                        customer_name, 'A', row_index
                    ),
                    'obj': self.validate_name(obj_name, 'B', row_index),
                    'doc_type': self.validate_doc_type(
                        doc_type, 'C', row_index
                    ),
                    'volume': self.validate_and_transform_volume(
                        volume, 'D', row_index
                    ),
                    'name': self.validate_name(name, 'E', row_index),
                    'filename': self.validate_name(filename, 'F', row_index),
                    'inventory_num': self.get_validate_inventory_num(
                        filename, 'F', row_index
                    )
                }
            )
            customers_names.add(customer_name)
            objects_names.add(obj_name)

        if self.errors:
            raise ExelParsingError

        return (albums, customers_names, objects_names)


class AlbumDataCreator:

    def __init__(
        self,
        album_data: tuple[list[dict[str, str | float], set[str], set[str]]]
    ) -> None:
        self.album_data = album_data

    def create_albums(self) -> int:
        albums, customers_names, objects_names = self.album_data
        with transaction.atomic():
            customers_map = self.handle_entries_with_name(
                Customer, customers_names
            )
            objects_map = self.handle_entries_with_name(Object, objects_names)

            for album_index in range(len(albums)):
                albums[album_index]['customer'] = customers_map[
                    albums[album_index]['customer']
                ]
                albums[album_index]['obj'] = objects_map[
                    albums[album_index]['obj']
                ]
                albums[album_index] = Album(**albums[album_index])

            Album.objects.bulk_create(albums)

        return len(albums)

    def handle_entries_with_name(
        self,
        model_class: Model,
        names_set: set
    ) -> dict[str, Model]:
        existing_entries = model_class.objects.filter(name__in=names_set)
        mapping_names = {entry.name: entry for entry in existing_entries}
        entries_to_create = [
                model_class(name=name) for name in names_set
                if name not in mapping_names
        ]
        if entries_to_create:
            created_entries = model_class.objects.bulk_create(
                entries_to_create
            )
            mapping_names.update(
                {entry.name: entry for entry in created_entries}
            )

        return mapping_names
